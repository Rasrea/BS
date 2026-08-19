"""
Excel 导出模块 - 生成4Sheet精简版家装报价Excel
"""
import json
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 样式
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
NORMAL_FONT = Font(name="微软雅黑", size=10)
MONEY_FONT = Font(name="微软雅黑", size=10, bold=True, color="C00000")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def export_quote_excel(quote_data: dict, output_dir: str = None) -> str:
    """
    生成4Sheet报价Excel
    
    参数:
        quote_data: 报价数据 dict，包含 items, totals 等
        output_dir: 输出目录，默认 ~/exports/
    
    返回:
        文件路径
    """
    if output_dir is None:
        output_dir = str(Path.home() / "exports")
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    project_name = quote_data.get("project_name", "智能报价单")
    timestamp = quote_data.get("create_time", "").replace(":", "-").replace(" ", "_")
    filename = f"{project_name}_{timestamp or 'export'}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    # ── Sheet1: 报价总表 ──
    ws1 = wb.active
    ws1.title = "报价总表"
    _build_summary_sheet(ws1, quote_data)

    # ── Sheet2: 分项明细 ──
    ws2 = wb.create_sheet("分项明细")
    _build_detail_sheet(ws2, quote_data)

    # ── Sheet3: 工程量清单 ──
    ws3 = wb.create_sheet("工程量清单")
    _build_quantity_sheet(ws3, quote_data)

    # ── Sheet4: 材质清单 ──
    ws4 = wb.create_sheet("材质清单")
    _build_material_sheet(ws4, quote_data)

    wb.save(filepath)
    return filepath


def _build_summary_sheet(ws, data: dict):
    """报价总表 Sheet"""
    ws.merge_cells("A1:F1")
    ws["A1"] = f"📋 {(data.get('project_name') or '家装智能报价单')}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:F2")
    ws["A2"] = f"生成时间: {data.get('create_time', '')}  |  规则版本: {data.get('rule_version', 'v1.0')}"
    ws["A2"].font = Font(name="微软雅黑", size=9, color="666666")
    ws["A2"].alignment = CENTER

    # 表头
    headers = ["费用项目", "金额(元)", "占比", "说明", "", ""]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    total = data.get("final_price", 0) or 1  # avoid div/0
    items = [
        ("基础工程量报价", data.get("base_price", 0), "CAD精准算量"),
        ("材质联动工序差价", data.get("material_diff_price", 0), "AI识别材质差价"),
        ("特殊工艺增项", data.get("process_add_price", 0), "造型/背景墙等"),
        ("工艺损耗费", data.get("loss_price", 0), "材料损耗"),
        ("管理费", data.get("manage_fee", 0), "项目管理"),
        ("税费", data.get("tax_fee", 0), "增值税"),
    ]

    row = 5
    for name, amount, note in items:
        ws.cell(row=row, column=1, value=name).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        c = ws.cell(row=row, column=2, value=round(amount, 2))
        c.font = MONEY_FONT
        c.border = THIN_BORDER
        c.alignment = RIGHT
        ratio = amount / total * 100
        ws.cell(row=row, column=3, value=f"{ratio:.1f}%").font = NORMAL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=4, value=note).font = NORMAL_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

    # 合计行
    row += 1
    cell = ws.cell(row=row, column=1, value="合 计")
    cell.font = Font(name="微软雅黑", size=12, bold=True, color="C00000")
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    c = ws.cell(row=row, column=2, value=round(total, 2))
    c.font = Font(name="微软雅黑", size=14, bold=True, color="C00000")
    c.alignment = RIGHT
    c.border = THIN_BORDER
    ws.cell(row=row, column=3, value="100%").font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=row, column=3).alignment = CENTER
    ws.cell(row=row, column=3).border = THIN_BORDER


def _build_detail_sheet(ws, data: dict):
    """分项明细 Sheet"""
    ws.merge_cells("A1:H1")
    ws["A1"] = "📋 分项报价明细"
    ws["A1"].font = TITLE_FONT

    headers = ["空间", "工程类别", "项目名称", "工程量", "单位", "材料单价", "人工单价", "小计"]
    widths = [14, 12, 16, 12, 8, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    items = data.get("items", [])
    row = 4
    for item in items:
        ws.cell(row=row, column=1, value=item.get("space_name", "")).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=item.get("category", "")).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=item.get("project_name", "")).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=item.get("quantity", 0)).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=item.get("unit", "㎡")).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=round(item.get("material_unit_price", 0), 2)).font = NORMAL_FONT
        ws.cell(row=row, column=7, value=round(item.get("labor_unit_price", 0), 2)).font = NORMAL_FONT
        c = ws.cell(row=row, column=8, value=round(item.get("subtotal", 0), 2))
        c.font = MONEY_FONT
        c.alignment = RIGHT
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = CENTER
        row += 1

    # 合计
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1, value="合 计").font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=1).border = THIN_BORDER
    total = sum(item.get("subtotal", 0) for item in items)
    c = ws.cell(row=row, column=8, value=round(total, 2))
    c.font = MONEY_FONT
    c.alignment = RIGHT
    c.border = THIN_BORDER


def _build_quantity_sheet(ws, data: dict):
    """工程量清单 Sheet"""
    ws.merge_cells("A1:F1")
    ws["A1"] = "📐 工程量清单（CAD源数据）"
    ws["A1"].font = TITLE_FONT

    headers = ["空间", "面积(㎡)", "周长(m)", "长(m)", "宽(m)", "高(m)"]
    widths = [16, 14, 12, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    cad_data = data.get("cad_data", [])
    row = 4
    for space in cad_data:
        ws.cell(row=row, column=1, value=space.get("space_name", space.get("name", ""))).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=round(space.get("area", 0), 2)).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=round(space.get("perimeter", 0), 2)).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=round(space.get("length", 0), 2)).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=round(space.get("width", 0), 2)).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=space.get("height", 2.8)).font = NORMAL_FONT
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = CENTER
        row += 1

    # 总面积
    if cad_data:
        row += 1
        ws.cell(row=row, column=1, value="总面积").font = Font(name="微软雅黑", size=11, bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        total_area = sum(s.get("area", 0) for s in cad_data)
        ws.cell(row=row, column=2, value=round(total_area, 2)).font = Font(name="微软雅黑", size=11, bold=True)
        ws.cell(row=row, column=2).border = THIN_BORDER


def _build_material_sheet(ws, data: dict):
    """材质清单 Sheet"""
    ws.merge_cells("A1:F1")
    ws["A1"] = "🎨 材质识别清单（AI识别结果）"
    ws["A1"].font = TITLE_FONT

    headers = ["空间", "墙面材质", "地面材质", "吊顶材质", "软装/其他", "置信度"]
    widths = [14, 16, 16, 14, 20, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    materials = data.get("material_data", [])
    row = 4
    for mat in materials:
        ws.cell(row=row, column=1, value=mat.get("space_name", "")).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=_get_mat(mat, "wall")).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=_get_mat(mat, "floor")).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=_get_mat(mat, "ceiling")).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=_get_mat(mat, "other")).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=f"{mat.get('confidence', 0)*100:.0f}%").font = NORMAL_FONT
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = CENTER
        row += 1


def _get_mat(mat: dict, key: str) -> str:
    """提取材质信息"""
    info = mat.get("material_info", {})
    if isinstance(info, str):
        return info[:30]
    if isinstance(info, dict):
        return info.get(key, info.get("空间描述", ""))[:30] or ""
    return str(info)[:30]


def export_process_quote_excel(quote_data: dict, output_dir: str = None) -> str:
    """
    新版精细化工序报价单导出
    格式：空间 → 墙/顶/地 → 细分施工工序 → 分级展示明细
    """
    if output_dir is None:
        output_dir = str(Path.home() / "exports")
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    project_name = quote_data.get("project_name", "精细化工序报价单")
    timestamp = quote_data.get("create_time", "").replace(":", "-").replace(" ", "_")
    filename = f"{project_name}_{timestamp or 'export'}_process.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    # ── Sheet1: 工序报价总表 ──
    ws1 = wb.active
    ws1.title = "工序报价总表"
    _build_process_summary(ws1, quote_data)

    # ── Sheet2: 空间→工序明细 ──
    ws2 = wb.create_sheet("空间工序明细")
    _build_space_process_detail(ws2, quote_data)

    # ── Sheet3: 分层工程量 ──
    ws3 = wb.create_sheet("分层工程量")
    _build_breakdown_detail(ws3, quote_data)

    # ── Sheet4: 材质与计价项 ──
    ws4 = wb.create_sheet("材质与计价项")
    _build_material_pricing_sheet(ws4, quote_data)

    wb.save(filepath)
    return filepath


def _build_process_summary(ws, data: dict):
    """Sheet1: 工序报价总表 - 按工序汇总"""
    ws.merge_cells("A1:G1")
    ws["A1"] = f"📋 {(data.get('project_name') or '精细化工序报价单')}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:G2")
    ws["A2"] = f"生成时间: {data.get('create_time', '')}  |  总分项: {len(data.get('cad_data', []))}个空间"
    ws["A2"].font = Font(name="微软雅黑", size=9, color="666666")
    ws["A2"].alignment = CENTER

    headers = ["工序名称", "工序类型", "工程量(㎡)", "人工费(元)", "材料费(元)", "辅料费(元)", "小计(元)"]
    widths = [16, 18, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    processes = data.get("processes", [])
    cad_data = data.get("cad_data", [])
    total_area = sum(r.get("area", 0) for r in cad_data) or 1

    row = 5
    total_labor = 0
    total_material = 0
    total_aux = 0
    total_all = 0

    for p in processes:
        pname = p.get("name", "")
        ptype = p.get("work_type", "")
        uprice = p.get("unit_price", 0)

        # 估算该工序覆盖的工程量（按空间面积比例简单分配）
        process_area = total_area * 0.8  # 默认80%覆盖率
        if ptype in ("demolition", "plumbing_electric"):
            process_area = total_area * 1.0
        elif ptype in ("waterproofing",):
            process_area = sum(r.get("area", 0) for r in cad_data
                               if any(kw in r.get("space_name", "") for kw in ["卫生间", "厨房", "阳台"])) or total_area * 0.3
        elif ptype in ("furnishing", "inspection"):
            process_area = 0

        labor = uprice * 0.4 * process_area
        material = uprice * 0.45 * process_area
        aux = uprice * 0.15 * process_area
        subtotal = labor + material + aux

        total_labor += labor
        total_material += material
        total_aux += aux
        total_all += subtotal

        ws.cell(row=row, column=1, value=pname).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=ptype).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=round(process_area, 2)).font = NORMAL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        for ci, val in [(4, labor), (5, material), (6, aux), (7, subtotal)]:
            c = ws.cell(row=row, column=ci, value=round(val, 2))
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = RIGHT
        row += 1

    # 总价行
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value="工序费合计").font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=1).border = THIN_BORDER
    for ci, val in [(4, total_labor), (5, total_material), (6, total_aux), (7, total_all)]:
        c = ws.cell(row=row, column=ci, value=round(val, 2))
        c.font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
        c.border = THIN_BORDER
        c.alignment = RIGHT

    # 总报价行
    row += 1
    final = data.get("final_price", 0)
    base = data.get("base_price", 0)
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="最终报价").font = Font(name="微软雅黑", size=12, bold=True, color="C00000")
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.cell(row=row, column=1).border = THIN_BORDER
    c = ws.cell(row=row, column=7, value=round(final, 2))
    c.font = Font(name="微软雅黑", size=14, bold=True, color="C00000")
    c.alignment = RIGHT
    c.border = THIN_BORDER


def _build_space_process_detail(ws, data: dict):
    """Sheet2: 空间→工序明细 - 按空间分墙/顶/地展示工序单价"""
    ws.merge_cells("A1:I1")
    ws["A1"] = "📋 空间→工序明细"
    ws["A1"].font = TITLE_FONT

    headers = ["空间名称", "面积(㎡)", "墙面工序", "墙面单价", "地面工序", "地面单价", "顶面工序", "顶面单价", "小计(元)"]
    widths = [14, 10, 16, 12, 16, 12, 16, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    cad_data = data.get("cad_data", [])
    pricing_items = data.get("pricing_items", [])
    bd_data = data.get("breakdown_data", [])

    # 建立材质→计价项映射
    # pricing_items按surface_type分组
    item_map = {}
    for pi in pricing_items:
        st = pi.get("surface_type", "wall")
        if st not in item_map:
            item_map[st] = []
        item_map[st].append(pi)

    row = 4
    grand_total = 0
    for space in cad_data:
        sname = space.get("space_name", "未命名")
        area = space.get("area", 0)
        detail = space.get("detail_json", {})
        if isinstance(detail, str):
            try:
                detail = json.loads(detail)
            except Exception:
                detail = {}
        sm = detail.get("surface_materials", {})

        # 根据surface_materials查找计价项
        wall_item = _find_pricing_item(item_map, "wall", sm.get("wall", {}).get("name", ""))
        floor_item = _find_pricing_item(item_map, "floor", sm.get("floor", {}).get("name", ""))
        ceiling_item = _find_pricing_item(item_map, "ceiling", sm.get("ceiling", {}).get("name", ""))

        wall_price = wall_item.get("unit_price", 65) if wall_item else 65
        floor_price = floor_item.get("unit_price", 85) if floor_item else 85
        ceiling_price = ceiling_item.get("unit_price", 35) if ceiling_item else 35

        # 估算各面面积
        wall_area = area * 2.5
        ceiling_area = area

        subtotal = wall_area * wall_price + area * floor_price + ceiling_area * ceiling_price
        grand_total += subtotal

        ws.cell(row=row, column=1, value=sname).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=round(area, 2)).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=wall_item.get("item_name", "乳胶漆墙面") if wall_item else "乳胶漆墙面").font = NORMAL_FONT
        ws.cell(row=row, column=4, value=wall_price).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=floor_item.get("item_name", "地砖铺贴") if floor_item else "地砖铺贴").font = NORMAL_FONT
        ws.cell(row=row, column=6, value=floor_price).font = NORMAL_FONT
        ws.cell(row=row, column=7, value=ceiling_item.get("item_name", "乳胶漆顶面") if ceiling_item else "乳胶漆顶面").font = NORMAL_FONT
        ws.cell(row=row, column=8, value=ceiling_price).font = NORMAL_FONT
        c = ws.cell(row=row, column=9, value=round(subtotal, 2))
        c.font = MONEY_FONT
        c.alignment = RIGHT

        for col in range(1, 10):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = CENTER
        row += 1

    # 合计
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="合 计").font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=1).border = THIN_BORDER
    c = ws.cell(row=row, column=9, value=round(grand_total, 2))
    c.font = Font(name="微软雅黑", size=12, bold=True, color="C00000")
    c.alignment = RIGHT
    c.border = THIN_BORDER


def _build_breakdown_detail(ws, data: dict):
    """Sheet3: 分层工程量"""
    ws.merge_cells("A1:J1")
    ws["A1"] = "📐 分层工程量明细"
    ws["A1"].font = TITLE_FONT

    headers = ["空间", "墙面面积(㎡)", "墙面单价", "地面面积(㎡)", "地面单价",
               "顶面面积(㎡)", "顶面单价", "墙面小计", "地面小计", "顶面小计"]
    widths = [12, 13, 11, 13, 11, 13, 11, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    cad_data = data.get("cad_data", [])
    pricing_items = data.get("pricing_items", [])
    item_map = {}
    for pi in pricing_items:
        st = pi.get("surface_type", "wall")
        if st not in item_map:
            item_map[st] = []
        item_map[st].append(pi)

    row = 4
    for space in cad_data:
        sname = space.get("space_name", "未命名")
        area = space.get("area", 0)
        detail = space.get("detail_json", {})
        if isinstance(detail, str):
            try:
                detail = json.loads(detail)
            except Exception:
                detail = {}
        sm = detail.get("surface_materials", {})

        wall_item = _find_pricing_item(item_map, "wall", sm.get("wall", {}).get("name", ""))
        floor_item = _find_pricing_item(item_map, "floor", sm.get("floor", {}).get("name", ""))
        ceiling_item = _find_pricing_item(item_map, "ceiling", sm.get("ceiling", {}).get("name", ""))

        wall_price = wall_item.get("unit_price", 65) if wall_item else 65
        floor_price = floor_item.get("unit_price", 85) if floor_item else 85
        ceiling_price = ceiling_item.get("unit_price", 35) if ceiling_item else 35

        wall_area = area * 2.5
        ceiling_area = area

        wall_sub = wall_area * wall_price
        floor_sub = area * floor_price
        ceiling_sub = ceiling_area * ceiling_price

        vals = [sname, round(wall_area, 2), wall_price, round(area, 2), floor_price,
                round(ceiling_area, 2), ceiling_price, round(wall_sub, 2),
                round(floor_sub, 2), round(ceiling_sub, 2)]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = CENTER
        row += 1


def _build_material_pricing_sheet(ws, data: dict):
    """Sheet4: 材质与计价项对照表"""
    ws.merge_cells("A1:G1")
    ws["A1"] = "🎯 材质 × 计价项对照"
    ws["A1"].font = TITLE_FONT

    headers = ["面类型", "计价项名称", "单位", "综合单价", "材料费", "人工费", "辅料费"]
    widths = [10, 16, 8, 12, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    pricing_items = data.get("pricing_items", [])
    surface_labels = {"wall": "墙面", "floor": "地面", "ceiling": "顶面", "all": "通用"}

    row = 4
    for pi in pricing_items:
        st = pi.get("surface_type", "")
        vals = [
            surface_labels.get(st, st),
            pi.get("item_name", ""),
            pi.get("unit", "㎡"),
            pi.get("unit_price", 0),
            pi.get("unit_price_material", 0),
            pi.get("unit_price_labor", 0),
            pi.get("unit_price_aux", 0),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=round(val, 2) if isinstance(val, float) else val)
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = CENTER
        row += 1


def _find_pricing_item(item_map: dict, surface_type: str, material_name: str) -> dict:
    """从计价项中找到匹配材质的最佳项"""
    items = item_map.get(surface_type, [])
    if not items:
        return {}
    if not material_name:
        return items[0]  # 默认第一个
    # 精确匹配
    for item in items:
        if material_name in item.get("item_name", "") or item.get("item_name", "") in material_name:
            return item
    # 子串匹配
    for item in items:
        iname = item.get("item_name", "")
        if any(kw in iname for kw in ["乳胶漆", "瓷砖", "墙纸", "地板", "木饰面", "大理石"]):
            if "乳胶漆" in material_name and "乳胶漆" in iname:
                return item
            if "瓷砖" in material_name and "瓷砖" in iname:
                return item
            if "地板" in material_name and "地板" in iname:
                return item
    return items[0]


# ─────────────────── 标准报价表导出 ───────────────────


def export_standard_report_excel(report_data: dict, output_dir: str = None) -> str:
    """
    导出标准报价表Excel（3个Sheet）
    
    Sheet1: 综合报价总表
    Sheet2: 空间分项明细表
    Sheet3: 工序费用明细表
    
    参数:
        report_data: standard_report API 返回的数据 dict
        output_dir: 输出目录，默认 ~/exports/
    
    返回:
        文件路径
    """
    if output_dir is None:
        output_dir = str(Path.home() / "exports")
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    project_name = report_data.get("project_name", "标准报价表")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"标准报价表_{project_name}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()

    # ── Sheet1: 综合报价总表 ──
    ws1 = wb.active
    ws1.title = "综合报价总表"
    _build_summary_sheet_from_report(ws1, report_data)

    # ── Sheet2: 空间分项明细表 ──
    ws2 = wb.create_sheet("空间分项明细")
    _build_space_detail_sheet(ws2, report_data)

    # ── Sheet3: 工序费用明细表 ──
    ws3 = wb.create_sheet("工序费用明细")
    _build_process_detail_sheet(ws3, report_data)

    wb.save(filepath)
    return filepath


def _build_summary_sheet_from_report(ws, data: dict):
    """Sheet1: 综合报价总表 - 项目概况 + 工种汇总"""
    # 标题
    ws.merge_cells("A1:E1")
    ws["A1"] = f"📋 {data.get('project_name', '标准报价表')}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    # 项目概况
    ws.merge_cells("A2:E2")
    ws["A2"] = f"报价编号: #{data.get('quote_id', '')}  |  生成时间: {data.get('create_time', '')}"
    ws["A2"].font = Font(name="微软雅黑", size=9, color="666666")
    ws["A2"].alignment = CENTER

    # 费用卡片区域
    row = 4
    headers = ["费用项目", "金额(元)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    fee_items = [
        ("基础报价", data.get("base_price", 0)),
        ("材质差价", data.get("material_diff", 0)),
        ("损耗", data.get("loss_price", 0)),
        ("管理费", data.get("manage_fee", 0)),
        ("税费", data.get("tax_fee", 0)),
        ("最终报价", data.get("total_price", 0)),
    ]
    row += 1
    for name, amount in fee_items:
        ws.cell(row=row, column=1, value=name).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = LEFT
        c = ws.cell(row=row, column=2, value=round(amount, 2))
        c.font = MONEY_FONT if name == "最终报价" else NORMAL_FONT
        c.border = THIN_BORDER
        c.alignment = RIGHT
        row += 1

    # 设置列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row=row, column=1, value="🔧 工种费用汇总").font = Font(name="微软雅黑", size=12, bold=True)
    row += 1

    summary_headers = ["工序", "空间数", "项目数", "金额(元)"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    row += 1

    process_summary = data.get("process_summary", [])
    for p in process_summary:
        ws.cell(row=row, column=1, value=p.get("process_name", "")).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=2, value=p.get("space_count", 0)).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=3, value=p.get("item_count", 0)).font = NORMAL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = CENTER
        c = ws.cell(row=row, column=4, value=round(p.get("subtotal", 0), 2))
        c.font = NORMAL_FONT
        c.border = THIN_BORDER
        c.alignment = RIGHT
        row += 1

    # 合计行
    ws.cell(row=row, column=1, value="合计").font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=1).alignment = LEFT
    total_spaces = sum(p.get("space_count", 0) for p in process_summary)
    total_items = sum(p.get("item_count", 0) for p in process_summary)
    total_amount = sum(p.get("subtotal", 0) for p in process_summary)
    ws.cell(row=row, column=2, value=total_spaces).font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).alignment = CENTER
    ws.cell(row=row, column=3, value=total_items).font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=row, column=3).border = THIN_BORDER
    ws.cell(row=row, column=3).alignment = CENTER
    c = ws.cell(row=row, column=4, value=round(total_amount, 2))
    c.font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
    c.border = THIN_BORDER
    c.alignment = RIGHT

    # 设置列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14


def _build_space_detail_sheet(ws, data: dict):
    """Sheet2: 空间分项明细表"""
    ws.merge_cells("A1:E1")
    ws["A1"] = "🏠 空间分项明细表"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    headers = ["空间", "项目名称", "数量", "单位", "小计(元)"]
    widths = [14, 20, 10, 8, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    space_details = data.get("space_details", [])
    row = 4
    for sd in space_details:
        space_name = sd.get("space_name", "")
        space_subtotal = sd.get("space_subtotal", 0)
        items = sd.get("items", [])

        for i, item in enumerate(items):
            ws.cell(row=row, column=1, value=space_name if i == 0 else "").font = NORMAL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=2, value=item.get("project_name", "")).font = NORMAL_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=2).alignment = LEFT
            ws.cell(row=row, column=3, value=item.get("quantity", 0)).font = NORMAL_FONT
            ws.cell(row=row, column=3).border = THIN_BORDER
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=4, value=item.get("unit", "")).font = NORMAL_FONT
            ws.cell(row=row, column=4).border = THIN_BORDER
            ws.cell(row=row, column=4).alignment = CENTER
            c = ws.cell(row=row, column=5, value=round(item.get("subtotal", 0), 2))
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = RIGHT
            row += 1

        # 空间小计行
        ws.cell(row=row, column=1, value=f"{space_name} 小计").font = Font(name="微软雅黑", size=10, bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = LEFT
        # 合并 B:D，但保留边框
        for c in range(2, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        ws.merge_cells(f"B{row}:D{row}")
        c = ws.cell(row=row, column=5, value=round(space_subtotal, 2))
        c.font = MONEY_FONT
        c.border = THIN_BORDER
        c.alignment = RIGHT
        row += 1

    # 总计行
    ws.merge_cells(f"A{row}:D{row}")
    total_cell = ws.cell(row=row, column=1, value="总计")
    total_cell.font = Font(name="微软雅黑", size=12, bold=True, color="C00000")
    total_cell.border = THIN_BORDER
    total_cell.alignment = LEFT
    # 设置整行边框
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = THIN_BORDER
    total = sum(sd.get("space_subtotal", 0) for sd in space_details)
    total_val = ws.cell(row=row, column=5, value=round(total, 2))
    total_val.font = Font(name="微软雅黑", size=12, bold=True, color="C00000")
    total_val.alignment = RIGHT
    total_val.border = THIN_BORDER


def _build_process_detail_sheet(ws, data: dict):
    """Sheet3: 工序费用明细表"""
    ws.merge_cells("A1:F1")
    ws["A1"] = "🔧 工序费用明细表"
    ws["A1"].font = TITLE_FONT

    headers = ["工序", "涉及空间", "空间数", "材料费", "人工费", "合计"]
    widths = [14, 24, 10, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    process_details = data.get("process_details", [])
    row = 4
    total_material = 0
    total_labor = 0
    total_subtotal = 0

    for p in process_details:
        ws.cell(row=row, column=1, value=p.get("process_name", "")).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=", ".join(p.get("spaces", []))).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=p.get("space_count", 0)).font = NORMAL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        mat = round(p.get("material_cost", 0), 2)
        lab = round(p.get("labor_cost", 0), 2)
        sub = round(p.get("subtotal", 0), 2)
        total_material += mat
        total_labor += lab
        total_subtotal += sub
        for ci, val in [(4, mat), (5, lab), (6, sub)]:
            c = ws.cell(row=row, column=ci, value=val)
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = RIGHT
        row += 1

    # 合计行
    ws.cell(row=row, column=1, value="合计").font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    total_spaces = len(set().union(*[set(p.get("spaces", [])) for p in process_details])) if process_details else 0
    ws.cell(row=row, column=2, value="").font = NORMAL_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3, value=total_spaces).font = Font(name="微软雅黑", size=11, bold=True)
    ws.cell(row=row, column=3).border = THIN_BORDER
    for ci, val in [(4, total_material), (5, total_labor), (6, total_subtotal)]:
        c = ws.cell(row=row, column=ci, value=round(val, 2))
        c.font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
        c.border = THIN_BORDER
        c.alignment = RIGHT
